"""
Generador de Etiquetas - ORO Construcción S.A.S
Aplicación web local para generar etiquetas en PDF (80x60mm landscape)
a partir de un archivo Excel con copias configurables por fila.
"""

import os
import io
import gc
import tempfile
import webbrowser
from datetime import datetime, date
from threading import Timer
from flask import Flask, render_template, request, send_file, jsonify
from openpyxl import load_workbook
from reportlab.lib.pagesizes import landscape
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from PIL import Image

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB máximo

# Dimensiones de la etiqueta: 80x60mm landscape
LABEL_WIDTH = 80 * mm
LABEL_HEIGHT = 60 * mm
PAGE_SIZE = (LABEL_WIDTH, LABEL_HEIGHT)

# Logo por defecto (ORO Construcción S.A.S)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_LOGO_PATH = os.path.join(BASE_DIR, 'logo_oro.jpeg')


def load_default_logo():
    """Carga el logo por defecto si existe."""
    if os.path.exists(DEFAULT_LOGO_PATH):
        with open(DEFAULT_LOGO_PATH, 'rb') as f:
            return f.read()
    return None

# Columnas esperadas en el Excel (normalizadas a minúsculas)
EXPECTED_COLUMNS = ['local', 'fecha', 'referencia', 'cod', 'precio', 'copias']


def normalize_column_name(name):
    """Normaliza el nombre de columna: minúsculas, sin espacios ni acentos."""
    if name is None:
        return ''
    name = str(name).strip().lower()
    # Eliminar acentos comunes
    replacements = {'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u', 'ñ': 'n'}
    for orig, repl in replacements.items():
        name = name.replace(orig, repl)
    return name


def read_excel_rows(excel_file):
    """
    Lee un archivo Excel y devuelve (rows, error_message).
    rows es una lista de diccionarios con las claves normalizadas.
    Si hay error, rows es None y error_message explica qué falló.
    """
    try:
        wb = load_workbook(excel_file, data_only=True, read_only=True)
        ws = wb.active
    except Exception as e:
        return None, f"No se pudo leer el Excel: {str(e)}"

    # Leer la primera fila como headers
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        return None, "El archivo Excel está vacío"

    headers = [normalize_column_name(h) for h in header_row]

    # Validar columnas requeridas
    missing = [col for col in EXPECTED_COLUMNS if col not in headers]
    if missing:
        wb.close()
        return None, (f"Faltan columnas en el Excel: {', '.join(missing)}. "
                      f"Se esperan: {', '.join(EXPECTED_COLUMNS)}")

    # Convertir cada fila en diccionario
    rows = []
    for row_values in rows_iter:
        # Saltar filas completamente vacías
        if all(v is None or (isinstance(v, str) and v.strip() == '') for v in row_values):
            continue
        row_dict = {}
        for i, header in enumerate(headers):
            if i < len(row_values):
                row_dict[header] = row_values[i]
            else:
                row_dict[header] = None
        rows.append(row_dict)

    wb.close()
    return rows, ""


def format_value(value):
    """Formatea un valor de celda para mostrarlo como texto limpio."""
    if value is None:
        return ''
    # Fechas de openpyxl vienen como datetime o date
    if isinstance(value, datetime):
        return value.strftime('%d/%m/%Y')
    if isinstance(value, date):
        return value.strftime('%d/%m/%Y')
    # Si es float pero entero (ej 123.0 -> 123)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def format_precio(value):
    """Formatea el precio con separador de miles y símbolo $."""
    if value is None:
        return ''
    try:
        num = float(value)
        if num.is_integer():
            # Formato colombiano: $ 1.234.567
            return f"$ {int(num):,}".replace(',', '.')
        else:
            return f"$ {num:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    except (ValueError, TypeError):
        return str(value)


def wrap_text_to_lines(text, max_width, font_name, font_size, canvas_obj, max_lines=2):
    """Divide texto en líneas que caben en max_width. Respeta palabras."""
    words = text.split()
    if not words:
        return ['']
    lines = []
    current = words[0]
    for word in words[1:]:
        test = current + ' ' + word
        if canvas_obj.stringWidth(test, font_name, font_size) <= max_width:
            current = test
        else:
            lines.append(current)
            current = word
            if len(lines) == max_lines - 1:
                # Última línea permitida, meter el resto
                remaining = ' '.join([current] + words[words.index(word) + 1:])
                lines.append(remaining)
                return lines
    lines.append(current)
    return lines


def fit_referencia(text, max_width, canvas_obj):
    """
    Encuentra el tamaño de fuente óptimo para la referencia,
    usando 1 o 2 líneas según haga falta.
    Devuelve (font_size, lines).
    """
    font_name = "Helvetica-Bold"
    # Probar tamaños desde grande a pequeño
    for font_size in range(22, 7, -1):
        # Intentar en 1 línea
        if canvas_obj.stringWidth(text, font_name, font_size) <= max_width:
            return font_size, [text]
        # Intentar en 2 líneas
        lines = wrap_text_to_lines(text, max_width, font_name, font_size, canvas_obj, max_lines=2)
        if len(lines) <= 2:
            max_line_width = max(canvas_obj.stringWidth(l, font_name, font_size) for l in lines)
            if max_line_width <= max_width:
                return font_size, lines
    # Fallback: tamaño mínimo, 2 líneas
    return 8, wrap_text_to_lines(text, max_width, font_name, 8, canvas_obj, max_lines=2)


def draw_label(c, data, logo_reader):
    """
    Dibuja una etiqueta completa en el canvas.
    Layout (80x60mm landscape):
      - Logo arriba-izquierda (más grande)
      - Fecha + Local arriba-derecha
      - Referencia centrada, grande, 1 o 2 líneas auto-ajustadas
      - Código debajo
      - Precio grande en la parte inferior
    """
    margin = 3 * mm
    w = LABEL_WIDTH
    h = LABEL_HEIGHT

    # === LOGO arriba-izquierda (más grande) ===
    logo_h_used = 0
    if logo_reader is not None:
        try:
            logo_max_w = 34 * mm
            logo_max_h = 16 * mm
            iw, ih = logo_reader.getSize()
            ratio = min(logo_max_w / iw, logo_max_h / ih)
            logo_w = iw * ratio
            logo_h = ih * ratio
            c.drawImage(
                logo_reader,
                margin,
                h - margin - logo_h,
                width=logo_w,
                height=logo_h,
                mask='auto',
                preserveAspectRatio=True,
            )
            logo_h_used = logo_h
        except Exception as e:
            print(f"Error dibujando logo: {e}")

    # === FECHA + LOCAL arriba-derecha ===
    fecha = format_value(data.get('fecha', ''))
    local = format_value(data.get('local', ''))
    c.setFont("Helvetica", 9)
    c.drawRightString(w - margin, h - margin - 7, fecha)
    c.setFont("Helvetica-Bold", 10)
    c.drawRightString(w - margin, h - margin - 17, f"Local: {local}")

    # === Definir zona para la referencia (entre header y código) ===
    # Header ocupa hasta ~18mm desde arriba
    header_bottom = h - margin - max(logo_h_used, 18 * mm) - 1 * mm
    # El precio+línea ocupan ~14mm desde abajo
    footer_top = margin + 14 * mm
    # Código ocupa ~7mm encima del precio
    cod_y = footer_top + 4 * mm
    # Zona disponible para referencia
    ref_zone_top = header_bottom
    ref_zone_bottom = cod_y + 5 * mm
    ref_zone_height = ref_zone_top - ref_zone_bottom

    # === REFERENCIA (auto-ajuste de tamaño, 1 o 2 líneas) ===
    referencia = format_value(data.get('referencia', ''))
    max_width = w - 2 * margin
    font_size, lines = fit_referencia(referencia, max_width, c)

    # Si caben 2 líneas pero la altura total es demasiada, reducir fuente
    line_height = font_size * 1.15
    total_ref_height = line_height * len(lines)
    while total_ref_height > ref_zone_height and font_size > 8:
        font_size -= 1
        _, lines = fit_referencia(referencia, max_width, c)
        # Recalcular con nueva fuente
        if len(lines) > 1:
            lines = wrap_text_to_lines(referencia, max_width, "Helvetica-Bold", font_size, c, max_lines=2)
        line_height = font_size * 1.15
        total_ref_height = line_height * len(lines)

    c.setFont("Helvetica-Bold", font_size)
    # Centrar verticalmente en la zona disponible
    ref_center_y = (ref_zone_top + ref_zone_bottom) / 2
    start_y = ref_center_y + (total_ref_height / 2) - font_size * 0.85
    for i, line in enumerate(lines):
        y = start_y - i * line_height
        c.drawCentredString(w / 2, y, line)

    # === CÓDIGO ===
    cod = format_value(data.get('cod', ''))
    c.setFont("Helvetica", 11)
    c.drawString(margin, cod_y, "Cód:")
    c.setFont("Helvetica-Bold", 12)
    c.drawString(margin + 12 * mm, cod_y, cod)

    # === Línea separadora ===
    c.setLineWidth(0.6)
    c.line(margin, footer_top - 1 * mm, w - margin, footer_top - 1 * mm)

    # === PRECIO (grande, destacado) ===
    precio = format_precio(data.get('precio', ''))
    # Ajustar tamaño si el precio es muy largo
    precio_font_size = 22
    while c.stringWidth(precio, "Helvetica-Bold", precio_font_size) > max_width and precio_font_size > 14:
        precio_font_size -= 1
    c.setFont("Helvetica-Bold", precio_font_size)
    c.drawCentredString(w / 2, margin + 4 * mm, precio)


def generate_pdf(rows, logo_bytes):
    """Genera el PDF con una etiqueta por página, repetida según 'copias'.
    rows: lista de diccionarios con las claves de EXPECTED_COLUMNS."""
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=PAGE_SIZE)

    # Preparar el logo
    logo_reader = None
    if logo_bytes:
        try:
            logo_img = Image.open(io.BytesIO(logo_bytes))
            # Convertir a RGBA para soportar transparencia
            if logo_img.mode != 'RGBA':
                logo_img = logo_img.convert('RGBA')
            logo_buffer = io.BytesIO()
            logo_img.save(logo_buffer, format='PNG')
            logo_buffer.seek(0)
            logo_reader = ImageReader(logo_buffer)
        except Exception as e:
            print(f"Error procesando logo: {e}")
            logo_reader = None

    total_labels = 0
    for row in rows:
        try:
            copias = int(float(row.get('copias', 1) or 1))
        except (ValueError, TypeError):
            copias = 1
        if copias < 1:
            copias = 1

        data = {col: row.get(col, '') for col in EXPECTED_COLUMNS}

        for _ in range(copias):
            draw_label(c, data, logo_reader)
            c.showPage()
            total_labels += 1

    c.save()
    buffer.seek(0)
    return buffer, total_labels


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/generar', methods=['POST'])
def generar():
    # Validar archivos
    if 'excel' not in request.files:
        return jsonify({'error': 'No se envió archivo Excel'}), 400

    excel_file = request.files['excel']
    if excel_file.filename == '':
        return jsonify({'error': 'Archivo Excel vacío'}), 400

    logo_bytes = None
    if 'logo' in request.files and request.files['logo'].filename != '':
        logo_bytes = request.files['logo'].read()
    else:
        logo_bytes = load_default_logo()

    # Leer y validar Excel
    rows, error = read_excel_rows(excel_file)
    if error:
        return jsonify({'error': error}), 400

    if not rows:
        return jsonify({'error': 'El Excel no tiene filas de datos'}), 400

    # Generar PDF
    try:
        pdf_buffer, total = generate_pdf(rows, logo_bytes)
    except Exception as e:
        return jsonify({'error': f'Error generando PDF: {str(e)}'}), 500
    finally:
        # Liberar explícitamente memoria
        rows = None
        logo_bytes = None
        gc.collect()

    return send_file(
        pdf_buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name='etiquetas_oro.pdf'
    )


@app.route('/vista-previa', methods=['POST'])
def vista_previa():
    """Genera solo la primera etiqueta como vista previa rápida."""
    if 'excel' not in request.files:
        return jsonify({'error': 'No se envió archivo Excel'}), 400

    excel_file = request.files['excel']
    logo_bytes = None
    if 'logo' in request.files and request.files['logo'].filename != '':
        logo_bytes = request.files['logo'].read()
    else:
        logo_bytes = load_default_logo()

    rows, error = read_excel_rows(excel_file)
    if error:
        return jsonify({'error': error}), 400

    if not rows:
        return jsonify({'error': 'El Excel no tiene filas de datos'}), 400

    # Solo primera fila, con copias forzadas a 1
    first_row = dict(rows[0])
    first_row['copias'] = 1

    try:
        pdf_buffer, _ = generate_pdf([first_row], logo_bytes)
    except Exception as e:
        return jsonify({'error': f'Error generando vista previa: {str(e)}'}), 500

    return send_file(
        pdf_buffer,
        mimetype='application/pdf',
        as_attachment=False,
        download_name='vista_previa.pdf'
    )


def find_free_port(preferred=5000):
    """
    Intenta usar el puerto preferido. Si está ocupado (app anterior colgada),
    intenta matar el proceso que lo ocupa. Si falla, usa otro puerto libre.
    """
    import socket
    import subprocess

    def is_port_in_use(port):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            return s.connect_ex(('127.0.0.1', port)) == 0

    # Si el puerto preferido está libre, úsalo
    if not is_port_in_use(preferred):
        return preferred

    # Puerto ocupado: intentar matar el proceso que lo ocupa (app anterior colgada)
    print(f"[AVISO] El puerto {preferred} está ocupado. Intentando liberarlo...")
    try:
        if os.name == 'nt':  # Windows
            # Buscar PID que usa el puerto
            result = subprocess.run(
                f'netstat -ano | findstr :{preferred}',
                shell=True, capture_output=True, text=True, timeout=5
            )
            pids = set()
            for line in result.stdout.strip().split('\n'):
                parts = line.split()
                if len(parts) >= 5 and f':{preferred}' in parts[1]:
                    pids.add(parts[-1])
            for pid in pids:
                if pid.isdigit() and pid != '0':
                    subprocess.run(
                        f'taskkill /F /PID {pid}',
                        shell=True, capture_output=True, timeout=5
                    )
                    print(f"[OK] Proceso {pid} cerrado.")
        else:  # Mac/Linux
            subprocess.run(
                f'lsof -ti:{preferred} | xargs kill -9',
                shell=True, capture_output=True, timeout=5
            )
        # Esperar un momento y verificar
        import time
        time.sleep(1)
        if not is_port_in_use(preferred):
            print(f"[OK] Puerto {preferred} liberado.")
            return preferred
    except Exception as e:
        print(f"[AVISO] No se pudo liberar el puerto: {e}")

    # Si sigue ocupado, buscar otro puerto libre
    for port in range(5001, 5020):
        if not is_port_in_use(port):
            print(f"[AVISO] Usando puerto alternativo: {port}")
            return port

    return preferred  # fallback (fallará, pero con mensaje claro)


def open_browser(port):
    webbrowser.open_new(f'http://127.0.0.1:{port}/')


if __name__ == '__main__':
    # Si hay PORT en variables de entorno (ej. Render), usarlo.
    # Si no, buscar puerto libre en local.
    env_port = os.environ.get('PORT')
    is_production = os.environ.get('RENDER') == 'true' or env_port is not None

    if is_production:
        port = int(env_port or 5000)
        app.run(host='0.0.0.0', port=port, debug=False)
    else:
        port = find_free_port(5000)
        print("=" * 60)
        print("  GENERADOR DE ETIQUETAS - ORO Construcción S.A.S")
        print("=" * 60)
        print(f"  Servidor corriendo en: http://127.0.0.1:{port}")
        print("  Presiona CTRL+C para detener el servidor")
        print("  (Si cierras esta ventana, la aplicación se detiene)")
        print("=" * 60)
        Timer(1.5, lambda: open_browser(port)).start()
        try:
            app.run(host='127.0.0.1', port=port, debug=False, use_reloader=False)
        except Exception as e:
            print(f"\n[ERROR] {e}")
            input("\nPresiona Enter para cerrar...")
